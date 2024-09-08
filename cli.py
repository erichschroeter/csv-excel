import argparse
import csv
import logging
import os
import textwrap
import yaml
from openpyxl import Workbook


def excel(args):
    config = None
    if args.config:
        with open(args.config, 'r') as yamlfile:
            config = yaml.safe_load(yamlfile)
            # for key, value in config.items():
            #     print(f"{key}: {value}")

    # Create a new workbook
    wb = Workbook()

    # Iterate over each CSV file
    for csv_file in args.csv_files:
        # Load the CSV file
        with open(csv_file, 'r') as f:
            reader = csv.reader(f)
            data = list(reader)

        # Create a new worksheet for this CSV file
        clean_title = os.path.basename(csv_file)  # don't inclue full path, just file name
        clean_title = os.path.splitext(clean_title)[0]  # remove extension
        sheet = wb.create_sheet(title=clean_title)

        if config:
            if clean_title in config['sheets']:
                sheet_config = config['sheets'][clean_title]
                if 'columns' in sheet_config:
                    for colname, colcfg in sheet_config['columns'].items():
                        if 'width' in colcfg:
                            logging.debug(f'Setting column "{colname}" to width of {colcfg["width"]}')
                            sheet.column_dimensions[colname].width = int(colcfg['width'])

        # Write the data to the worksheet
        for row in data:
            sheet.append(row)

    # Delete the default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    # Save the workbook
    wb.save('output.xlsx')


class App:
    def __init__(self) -> None:
        self.args = None
        self.parser = argparse.ArgumentParser(
            description=textwrap.dedent('''\
                A commandline utility to manage an Excel file with multiple worksheets while keeping data in CSV files for better Git support.
                '''),
            formatter_class=RawTextArgumentDefaultsHelpFormatter)
        self.parser.add_argument('-v', '--verbosity',
                                 choices=['critical', 'error', 'warning', 'info', 'debug'],
                                 default='info',
                                 help='Set the logging verbosity level.')
        self.parser.add_argument('-c', '--config',
                                 help='A YAML configuration file.')

        self.subparsers = self.parser.add_subparsers(dest='command')
        excel_parser = self.subparsers.add_parser('excel',
                                                   help='Generate or update an Excel file from multiple CSV files.',
                                                   formatter_class=RawTextArgumentDefaultsHelpFormatter)
        excel_parser.add_argument('csv_files', nargs='+', help='The CSV files to include in the Excel file')
        excel_parser.set_defaults(func=excel)

    def parse_args(self, args=None):
        self.args = self.parser.parse_args(args)

    def run(self):
        if not self.args:
            self.parse_args()
        _init_logger(getattr(logging, self.args.verbosity.upper()))
        logging.debug(f'command-line args: {self.args}')
        self.args.func(self.args)


class ColorLogFormatter(logging.Formatter):
    '''
    Custom formatter that changes the color of logs based on the log level.
    '''

    grey = "\x1b[38;20m"
    green = "\u001b[32m"
    yellow = "\x1b[33;20m"
    red = "\x1b[31;20m"
    bold_red = "\x1b[31;1m"
    blue = "\u001b[34m"
    cyan = "\u001b[36m"
    reset = "\x1b[0m"

    timestamp = '%(asctime)s - '
    loglevel = '%(levelname)s'
    message = ' - %(message)s'

    FORMATS = {
        logging.DEBUG:    timestamp + blue + loglevel + reset + message,
        logging.INFO:     timestamp + green + loglevel + reset + message,
        logging.WARNING:  timestamp + yellow + loglevel + reset + message,
        logging.ERROR:    timestamp + red + loglevel + reset + message,
        logging.CRITICAL: timestamp + bold_red + loglevel + reset + message
    }

    def format(self, record):
        log_fmt = self.FORMATS.get(record.levelno)
        formatter = logging.Formatter(log_fmt)
        return formatter.format(record)


def _init_logger(level=logging.INFO):
    logger = logging.getLogger()
    logger.setLevel(level)

    formatter = ColorLogFormatter()
    # create console handler and set level to debug
    ch = logging.StreamHandler()
    ch.setLevel(level)
    ch.setFormatter(formatter)
    logger.addHandler(ch)


class RawTextArgumentDefaultsHelpFormatter(argparse.ArgumentDefaultsHelpFormatter, argparse.RawTextHelpFormatter):
    pass


if __name__ == '__main__':
    App().run()
