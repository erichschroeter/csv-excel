import argparse
import csv
import importlib
import logging
import openpyxl
import os
import sys
import textwrap
from pathlib import Path
import xlsxwriter
from xlsxwriter.utility import xl_cell_to_rowcol
import yaml


def column_to_index(col_str):
    """
    Convert a column cell reference notation to a zero indexed row and column.
    For example, 'A' will assume 'A1' and return (0,0)

    Args:
       col_str:  The column for A1 style string.

    Returns:
        row, col: Zero indexed cell row and column indices.

    """
    return xl_cell_to_rowcol(f'{col_str.upper()}1')[0]


class WorkbookFactory:
    def __init__(self, config_path=None) -> None:
        self.config_path = config_path
        if self.config_path:
            with open(self.config_path, 'r') as yamlfile:
                self.config = yaml.safe_load(yamlfile)

    def _csv_path_to_worksheet_title(self, csv_path) -> str:
        title = os.path.basename(csv_path)  # don't include full path, just file name
        title = os.path.splitext(title)[0]  # remove extension
        return title

    def build_openpyxl(self, output_path, csv_files):
        pass

    def build_xslxwriter(self, output_path, csv_files):
        wb = xlsxwriter.Workbook(output_path)
        # Delete the default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        # Include the Excel macro that auto exports worksheets to CSV files when file is saved.
        vbaproject_path = f'{os.path.dirname(os.path.abspath(__file__))}/vbaProject.bin'
        logging.debug(f'Packing VBA project into Excel file: "{vbaproject_path}"')
        wb.add_vba_project(vbaproject_path)

        for csv_file in csv_files:
            with open(csv_file, 'r') as f:
                reader = csv.reader(f)
                csv_data = list(reader)

            worksheet_title = self._csv_path_to_worksheet_title(csv_file)
            sheet = wb.add_worksheet(name=worksheet_title)
            logging.debug(f'Added worksheet "{worksheet_title}"')

            # Apply any config specifications.
            if self.config:
                if worksheet_title in self.config['sheets']:
                    sheet_config = self.config['sheets'][worksheet_title]
                    if 'columns' in sheet_config:
                        for colname, colcfg in sheet_config['columns'].items():
                            if 'width' in colcfg:
                                width = int(colcfg['width'])
                                colindex = column_to_index(colname)
                                sheet.set_column_pixels(colindex, colindex, width)

            # Write the data to the worksheet
            for row, data in enumerate(csv_data):
                sheet.write_row(row, 0, data)

        return wb


def csv2xl(args):
    """
    Generates or updates an Excel file from multiple CSV files.

    Args:
        args:  The command line args.
    """
    wb = WorkbookFactory(args.config).build_xslxwriter(args.output, args.csv_files)
    # Save the workbook
    wb.close()


def xl2csv(args):
    """
    Exports worksheets within an Excel file to CSV files.

    Args:
        args:  The command line args.
    """
    wb = openpyxl.load_workbook(args.file)
    # Create the output directory if it does not exist.
    if args.output_dir:
        Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    for sheet in wb:
        with open(os.path.join(args.output_dir if args.output_dir else '', f'{sheet.title}.csv'), 'w+', newline='') as f:
            logging.debug(f'Exporting worksheet "{sheet.title}"')
            c = csv.writer(f)
            for row in sheet.rows:
                c.writerow([cell.value for cell in row])


def validate(args):
    wb = WorkbookFactory(args.config).build_openpyxl('output.xlsm', args.csv_files)

    from os.path import dirname, basename, isfile, join
    import glob
    modules = glob.glob(join(args.rules_dir, '*.py'))
    # modules = [join('csv_excel.rules.', m) for m in glob.glob(join(dirname(__file__), 'rules/*.py'))]
    logging.warning(f'Found modules: {modules}')
    # rule_modules = [ basename(f)[:-3] for f in modules if isfile(f) and not f.endswith('__init__.py')]
    rule_modules = [ f'csv_excel.rules.{basename(f)[:-3]}' for f in modules if isfile(f) and not f.endswith('__init__.py')]
    logging.warning(f'Checking rules: {rule_modules}')
    for rule_module in rule_modules:
        # rule = importlib.import_module('csv_excel.rules.unique_id')
        rule = importlib.import_module(rule_module)
        v = getattr(rule, 'validate')
        v(wb)


def dir_path(string):
    if os.path.isdir(string):
        return string
    else:
        raise NotADirectoryError(string)


class App:
    def __init__(self) -> None:
        self.args = None
        self.parser = argparse.ArgumentParser(
            description=textwrap.dedent('''\
                A commandline utility to manage an Excel file with multiple worksheets while keeping data in CSV files for better Git support.
                '''),
            formatter_class=RawTextArgumentDefaultsHelpFormatter)
        self.parser.add_argument('-v', '--verbosity',
                                 choices=['critical', 'error', 'warning', 'info', 'debug'],
                                 default='info',
                                 help='Set the logging verbosity level.')
        self.parser.add_argument('-c', '--config',
                                 help='A YAML configuration file.')

        self.subparsers = self.parser.add_subparsers(dest='command')
        csv2xl_parser = self.subparsers.add_parser('csv2xl',
                                                   help='Generate or update an Excel file from multiple CSV files.',
                                                   formatter_class=RawTextArgumentDefaultsHelpFormatter)
        csv2xl_parser.add_argument('-o', '--output', default='output.xlsm', help='The output Excel file')
        csv2xl_parser.add_argument('csv_files', nargs='+', help='The CSV files to include in the Excel file')
        csv2xl_parser.set_defaults(func=csv2xl)
        xl2csv_parser = self.subparsers.add_parser('xl2csv',
                                                   help='Exports worksheets to CSV files.',
                                                   formatter_class=RawTextArgumentDefaultsHelpFormatter)
        xl2csv_parser.add_argument('-o', '--output_dir', help='The output Excel file')
        xl2csv_parser.add_argument('file', help='The Excel file')
        xl2csv_parser.set_defaults(func=xl2csv)
        validate_parser = self.subparsers.add_parser('validate',
                                                     help='Validate a set of CSV files given a set of rules.',
                                                     formatter_class=RawTextArgumentDefaultsHelpFormatter)
        validate_parser.add_argument('csv_files', nargs='+', help='The CSV files to include in the Excel file')
        validate_parser.add_argument('--rules_dir',
                                     type=dir_path,
                                     default=os.path.join(os.getcwd(), 'rules'),
                                     help='Directory path to rules')
        validate_parser.set_defaults(func=validate)

    def parse_args(self, args=None):
        self.args = self.parser.parse_args(args)

    def run(self):
        if not self.args:
            self.parse_args()
        # try:
        #     if not self.args:
        #         self.parse_args()
        # except:
        #     self.parser.print_help()
        #     sys.exit(1)
        _init_logger(getattr(logging, self.args.verbosity.upper()))
        logging.debug(f'command-line args: {self.args}')
        self.args.func(self.args)


class ColorLogFormatter(logging.Formatter):
    '''
    Custom formatter that changes the color of logs based on the log level.
    '''

    grey = "\x1b[38;20m"
    green = "\u001b[32m"
    yellow = "\x1b[33;20m"
    red = "\x1b[31;20m"
    bold_red = "\x1b[31;1m"
    blue = "\u001b[34m"
    cyan = "\u001b[36m"
    reset = "\x1b[0m"

    timestamp = '%(asctime)s - '
    loglevel = '%(levelname)s'
    message = ' - %(message)s'

    FORMATS = {
        logging.DEBUG:    timestamp + blue + loglevel + reset + message,
        logging.INFO:     timestamp + green + loglevel + reset + message,
        logging.WARNING:  timestamp + yellow + loglevel + reset + message,
        logging.ERROR:    timestamp + red + loglevel + reset + message,
        logging.CRITICAL: timestamp + bold_red + loglevel + reset + message
    }

    def format(self, record):
        log_fmt = self.FORMATS.get(record.levelno)
        formatter = logging.Formatter(log_fmt)
        return formatter.format(record)


def _init_logger(level=logging.INFO):
    logger = logging.getLogger()
    logger.setLevel(level)

    formatter = ColorLogFormatter()
    # create console handler and set level to debug
    ch = logging.StreamHandler()
    ch.setLevel(level)
    ch.setFormatter(formatter)
    logger.addHandler(ch)


class RawTextArgumentDefaultsHelpFormatter(argparse.ArgumentDefaultsHelpFormatter, argparse.RawTextHelpFormatter):
    pass


if __name__ == '__main__':
    App().run()
