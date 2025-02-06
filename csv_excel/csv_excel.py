import csv
import glob
import importlib
import importlib.util
import inspect
import logging
import openpyxl
import os
from os.path import dirname, basename, isfile, join
from pathlib import Path
import xlsxwriter
from xlsxwriter.utility import xl_cell_to_rowcol
import yaml


def column_to_index(col_str):
    """
    Convert a column cell reference notation to a zero indexed row and column.
    For example, 'A' will assume 'A1' and return (0,0)

    Args:
       col_str:  The column for A1 style string.

    Returns:
        row, col: Zero indexed cell row and column indices.

    """
    return xl_cell_to_rowcol(f"{col_str.upper()}1")[1]


class RuleError(Exception):
    def __init__(self, rule_file, message, *args: object) -> None:
        self.message = f"{os.path.basename(rule_file)}: {message}"
        super().__init__(message, *args)


class CsvRuleError(RuleError):
    def __init__(self, rule_file, row, col, message, *args: object) -> None:
        super().__init__(rule_file, f"(row: {row}, col: {col}): {message}", *args)


class CsvSheet:
    def __init__(self, path=None, data_row=1, freeze_pane_row=0, freeze_pane_col=0):
        self.path = path
        self.data_row = data_row
        self.freeze_pane_row = freeze_pane_row
        self.freeze_pane_col = freeze_pane_col
        self.reader = None
        self._data = ["A", "B", "C", "D", "E"]

    def data(self):
        for row in self._data:
            yield row
        # if self.path:
        #     with open(self.path, newline="") as csvfile:
        #         csv_basename = os.path.basename(self.path).split(".")[0]
        #         r = csv.reader(csvfile)
        #         rownum = 0
        #         if self.config and "sheets" in self.config:
        #             if csv_basename in self.config["sheets"]:
        #                 if "data_row" in self.config["sheets"][csv_basename]:
        #                     for _ in range(
        #                         self.config["sheets"][csv_basename]["data_row"] - 1
        #                     ):
        #                         next(r)
        #                         rownum += 1
        #         for i, row in enumerate(r):
        #             rownum += 1
        #             if i >= n:
        #                 break
        #             yield self.path, csv_basename, rownum, row


class WorkbookFactory:
    def __init__(self, config=None, csv_files=[]) -> None:
        # To support testability, provide a way to override config handlers.
        self.handlers = {
            "set_column_width": self._set_column_width,
        }
        self.config_path = None
        self.config = config
        self.csv_data_readers = csv_files

    def with_config(self, config_path):
        self.config_path = config_path
        with open(self.config_path, "r") as yamlfile:
            logging.debug(f"Loading config: {self.config_path}")
            self.config = yaml.safe_load(yamlfile)
        return self

    def with_csv_files(self, csv_file_paths):
        return self

    def _csv_path_to_worksheet_title(self, csv_path) -> str:
        title = os.path.basename(csv_path)  # don't include full path, just file name
        title = os.path.splitext(title)[0]  # remove extension
        return title

    def _set_column_width(self, sheet, column_name, width):
        colindex = column_to_index(column_name)
        logging.debug(
            f'Sheet "{sheet.get_name()}" column "{column_name}" ({colindex}) to {width}px'
        )
        sheet.set_column_pixels(colindex, colindex, width)

    def build_openpyxl(self, csv_files, output_path=None):
        wb = openpyxl.Workbook()
        # Delete the default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        for csv_file in csv_files:
            with open(csv_file, "r") as f:
                reader = csv.reader(f)
                csv_data = list(reader)

            worksheet_title = self._csv_path_to_worksheet_title(csv_file)
            sheet = wb.create_sheet(title=worksheet_title)
            logging.debug(f'Added worksheet "{worksheet_title}"')

            # Write the data to the worksheet
            for data in csv_data:
                sheet.append(data)
        if output_path:
            wb.save(output_path)
        return wb

    def build_xlsxwriter(self, csv_files, output_path):
        wb = xlsxwriter.Workbook(output_path)
        # Delete the default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        # Include the Excel macro that auto exports worksheets to CSV files when file is saved.
        vbaproject_path = f"{os.path.dirname(os.path.abspath(__file__))}/vbaProject.bin"
        logging.debug(f'Packing VBA project into Excel file: "{vbaproject_path}"')
        wb.add_vba_project(vbaproject_path)

        for csv_file in csv_files:
            with open(csv_file, "r") as f:
                reader = csv.reader(f)
                csv_data = list(reader)

            worksheet_title = self._csv_path_to_worksheet_title(csv_file)
            sheet = wb.add_worksheet(name=worksheet_title)
            logging.debug(f'Added worksheet "{worksheet_title}"')

            # Apply any config specifications.
            if self.config:
                if worksheet_title in self.config["sheets"]:
                    sheet_config = self.config["sheets"][worksheet_title]
                    if "columns" in sheet_config:
                        for colname, colcfg in sheet_config["columns"].items():
                            if "width" in colcfg:
                                self.handlers["set_column_width"](
                                    sheet, colname, int(colcfg["width"])
                                )

            # Write the data to the worksheet
            for row, data in enumerate(csv_data):
                sheet.write_row(row, 0, data)

        return wb


def validate_data_isolated(file_path):
    pass


def validate_data_correlated(file_path):
    pass


def csv2xl(args):
    """
    Generates or updates an Excel file from multiple CSV files.

    Args:
        args:  The command line args.
    """
    # Use xlsxwriter due to support for vbaProject macros.
    wb = WorkbookFactory(args.config).build_xlsxwriter(args.csv_files, args.output)
    # Save the workbook
    wb.close()


def xl2csv(args):
    """
    Exports worksheets within an Excel file to CSV files.

    Args:
        args:  The command line args.
    """
    wb = openpyxl.load_workbook(args.file)
    # Create the output directory if it does not exist.
    if args.output_dir:
        Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    for sheet in wb:
        with open(
            os.path.join(
                args.output_dir if args.output_dir else "", f"{sheet.title}.csv"
            ),
            "w+",
            newline="",
            encoding="utf-8",
        ) as f:
            logging.debug(f'Exporting worksheet "{sheet.title}"')
            c = csv.writer(f)
            for row in sheet.rows:
                c.writerow([cell.value for cell in row])


def collect_csv_data_rules(file_path):
    """
    Reads the file for functions annotated with @csv_data_rule.
    """
    # Load the modules from the file path
    spec = importlib.util.spec_from_file_location("temp_module", file_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    # Collect all the functions annotated with @csv_data_rule
    rules = []
    for _, obj in inspect.getmembers(module, inspect.isfunction):
        if hasattr(obj, "_is_csv_data_rule") and obj._is_csv_data_rule:
            logging.debug(f"Found csv_data_rule in {file_path}: {obj.__name__}")
            rules.append(obj)
    return rules


def collect_workbook_rules(file_path):
    """
    Reads the file for functions annotated with @workbook_rule.
    """
    # Load the modules from the file path
    spec = importlib.util.spec_from_file_location("temp_module", file_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    # Collect all the functions annotated with @workbook_rule
    rules = []
    for _, obj in inspect.getmembers(module, inspect.isfunction):
        if hasattr(obj, "_is_workbook_rule") and obj._is_workbook_rule:
            logging.debug(f"Found workbook rule in {file_path}: {obj.__name__}")
            rules.append(obj)
    return rules


def csv_data_rule(*args, **kwargs):
    applies_to = kwargs.get("applies_to", None)

    if len(args) == 1 and callable(args[0]):
        func = args[0]
        func._is_csv_data_rule = True
        func._applies_to = applies_to
        return func

    def decorator(func):
        func._is_csv_data_rule = True
        func._applies_to = applies_to
        return func

    return decorator


def workbook_rule(*args, **kwargs):
    if len(args) == 1 and callable(args[0]):
        func = args[0]
        func._is_workbook_rule = True
        return func

    def decorator(func):
        func._is_workbook_rule = True
        return func

    return decorator


def directory_to_module_path(directory_path):
    # Normalize the path to use the correct OS-specific separator
    normalized_path = os.path.normpath(directory_path)

    # Split the path into components
    path_components = normalized_path.split(os.sep)

    # Remove the file extension if present
    if path_components[-1].endswith(".py"):
        path_components[-1] = path_components[-1][:-3]

    # Join the components with dots to form the module path
    module_path = ".".join(path_components)

    return module_path


def validate(args):
    if not args.rules:
        return

    modules = []
    for path in args.rules:
        if os.path.isdir(path):
            globbed = sorted(glob.glob(join(path, "*.py")))
            modules.extend(globbed)
        elif os.path.isfile(path):
            modules.append(path)
    modules = sorted(modules)
    logging.debug(f"Found modules: {modules}")

    rules = []
    # Validate CSV data rules.
    for module_path in modules:
        rules.extend(collect_csv_data_rules(module_path))
    for file_path in args.csv_files:
        for rule in rules:
            if rule._applies_to is None or basename(file_path) in rule._applies_to:
                with open(file_path, newline="") as f:
                    reader = csv.DictReader(f)
                    for row_num, row in enumerate(reader):
                        try:
                            errors = rule(row, row_num)
                        except RuleError as e:
                            logging.error(f"{file_path}: {e.message}")

    # Validate workbook rules.
    # Use openpyxl due to better support for reading data.
    wb = WorkbookFactory(args.config).build_openpyxl(args.csv_files)
    rules = []
    for module_path in modules:
        rules.extend(collect_workbook_rules(module_path))
    for rule in rules:
        try:
            errors = rule(wb)
        except RuleError as e:
            logging.error(e.message)

    #     v = getattr(rule, "validate")
    #     result = v(wb)
    #     if result:
    #         results.extend(result)
    # if results:
    #     for result in results:
    #         logging.error(f"{result.message}")
